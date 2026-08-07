"""TestHub Agent 业务工具（OpenAI Agents SDK 原生 function tool）。

每个工具是带类型注解的普通函数，经 @assistant_tool 注册后自动进入
SDK 工具集。权限校验在工具体内部显式调用 require_* 完成；任何
ToolPermissionError 或未捕获异常都会通过 failure_error_function
转成结构化错误返回给模型，而不是中断整个运行。
"""
from __future__ import annotations

import ipaddress
import json
import logging
import os
import shlex
import shutil
import subprocess
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

from agents import FunctionTool, RunContextWrapper, function_tool
from agents.exceptions import UserError

from apps.assistant.context import TestHubContext

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------
# 注册表与权限基础
# ---------------------------------------------------------------

TOOL_REGISTRY: Dict[str, FunctionTool] = {}

# 工具按后台功能模块分组，用于按场景裁剪，控制模型上下文占用。
# tool_groups 为空表示启用全部工具（向后兼容）。
TOOL_GROUPS: Dict[str, List[str]] = {
    "project": [
        "get_project_overview",
        "update_knowledge_base",
        "read_knowledge_base",
    ],
    "api_testing": [
        "search_apis",
        "get_api_detail",
        "create_api_test",
        "update_api_test",
        "create_collection",
        "execute_api",
        "list_api_projects",
        "create_test_suite",
        "list_test_suites",
        "add_suite_request",
        "list_suite_requests",
        "update_suite_request",
        "delete_suite_request",
        "execute_test_suite",
    ],
    "testcases": [
        "search_testcases",
        "analyze_testcases",
        "create_testcase",
        "get_testcase_detail",
        "update_testcase",
        "delete_testcase",
    ],
    "ui_automation": [
        "list_midscene_projects",
        "list_midscene_cases",
        "update_midscene_case",
    ],
    "documents": [
        "parse_swagger",
        "parse_yapi",
        "list_session_files",
        "read_session_file",
        "simple_doc_parser",
    ],
    "browser": [
        "agent_browser",
    ],
}


def resolve_tool_names(tool_groups: Optional[List[str]] = None) -> List[str]:
    """按工具组解析要启用的工具名；空/None 表示全部工具。"""
    if not tool_groups:
        return list(TOOL_REGISTRY.keys())
    names = []
    for group in tool_groups:
        names.extend(TOOL_GROUPS.get(group, []))
    if not names:
        # 配置的组全部未知时回退全部，避免 agent 失去所有工具
        return list(TOOL_REGISTRY.keys())
    # 去重并保持注册顺序
    return [n for n in TOOL_REGISTRY if n in names]


def _slim_json(value: Any, limit: int = 2000) -> tuple:
    """将任意值序列化为 JSON 字符串并截断，返回 (截断后文本, 是否被截断)。

    用于限制工具返回进入模型上下文的体积；仅影响返回内容，
    不修改数据库中的原始数据。
    """
    try:
        s = json.dumps(value, ensure_ascii=False, indent=2, default=str)
    except Exception:
        s = str(value)
    return s[:limit], len(s) > limit


class ToolPermissionError(Exception):
    """工具权限/参数校验失败。"""


def _require(cond: bool, msg: str) -> None:
    if not cond:
        raise ToolPermissionError(msg)


def _user_has_project_access(user: Any, project_id: Optional[int]) -> bool:
    """判断 user 是否为项目成员/负责人（超管放行）。"""
    if user is None or not project_id:
        return False
    if getattr(user, "is_superuser", False):
        return True
    from apps.projects.models import Project, ProjectMember

    return ProjectMember.objects.filter(
        project_id=project_id, user=user
    ).exists() or Project.objects.filter(id=project_id, owner=user).exists()


def require_project(ctx: RunContextWrapper[TestHubContext], project_id: Optional[int]) -> None:
    """校验当前用户对主项目（Project）的访问权。"""
    _require(project_id, "缺少 project_id 参数")
    if not _user_has_project_access(ctx.context.user, project_id):
        raise ToolPermissionError(f"当前用户无权访问项目 {project_id}")


def _module_access(user: Any, obj: Any) -> bool:
    """模块级项目（ApiProject/MidsceneProject 等）的 owner/members 校验。"""
    if user is None:
        return False
    if getattr(user, "is_superuser", False):
        return True
    if getattr(obj, "owner_id", None) == user.id:
        return True
    members = getattr(obj, "members", None)
    return bool(members and members.filter(id=user.id).exists())


def _resolve_api_request(request_id: int) -> tuple:
    from apps.api_testing.models import ApiRequest

    req = ApiRequest.objects.select_related("collection", "collection__project").get(
        id=request_id
    )
    api_project = req.collection.project if req.collection and req.collection.project else None
    return req, api_project


def require_api_access(ctx: RunContextWrapper[TestHubContext], request_id: int) -> Any:
    """解析接口并校验访问权（主项目优先，其次模块 owner/members）。"""
    req, api_project = _resolve_api_request(request_id)
    if api_project is None:
        _require(
            getattr(ctx.context.user, "is_superuser", False),
            "无权操作未归属项目的接口",
        )
        return req
    if api_project.main_project_id:
        require_project(ctx, api_project.main_project_id)
    else:
        _require(
            _module_access(ctx.context.user, api_project),
            f"无权访问接口项目 {api_project.name}",
        )
    return req


def require_midscene_access(
    ctx: RunContextWrapper[TestHubContext], midscene_project_id: int
) -> Any:
    """校验 Midscene 项目访问权。"""
    from apps.ui_automation.models import MidsceneProject

    mp = MidsceneProject.objects.get(id=midscene_project_id)
    if mp.main_project_id:
        require_project(ctx, mp.main_project_id)
    else:
        _require(_module_access(ctx.context.user, mp), f"无权访问 Midscene 项目 {mp.name}")
    return mp


def require_testcase_access(
    ctx: RunContextWrapper[TestHubContext], testcase_id: int
) -> Any:
    """解析测试用例并校验其所属项目访问权。"""
    from apps.testcases.models import TestCase

    tc = TestCase.objects.get(id=testcase_id)
    require_project(ctx, tc.project_id)
    return tc


def _user_session_ids(ctx: RunContextWrapper[TestHubContext]) -> List[str]:
    from apps.assistant.models import AssistantSession

    user = ctx.context.user
    if user is None:
        return []
    return list(
        AssistantSession.objects.filter(user=user).values_list("session_id", flat=True)
    )


def _safe_session_path(ctx: RunContextWrapper[TestHubContext], fp: str) -> str:
    """限定文件读取范围为当前用户的会话上传目录。"""
    from django.conf import settings

    root = os.path.realpath(os.path.join(settings.MEDIA_ROOT, "uploads", "agent"))
    resolved = os.path.realpath(fp)
    if not resolved.startswith(root + os.sep):
        raise ToolPermissionError("只能读取会话上传目录内的文件")
    allowed = {os.path.join(root, sid) for sid in _user_session_ids(ctx)}
    if not any(resolved.startswith(d + os.sep) for d in allowed):
        raise ToolPermissionError("只能读取当前用户会话目录内的文件")
    return resolved


def _validate_target_url(url: str, environment: Any = None) -> str:
    """拦截内网/本机地址（SSRF 基础防护），返回规整后的 URL。"""
    if environment and getattr(environment, "base_url", None):
        url = url.replace("{{base_url}}", environment.base_url.rstrip("/"))
    if "{{" in url:
        return url  # 含未解析模板时交由执行层处理
    parsed = urlparse(url)
    _require(parsed.scheme in ("http", "https"), f"不允许的协议: {parsed.scheme}")
    host = parsed.hostname
    _require(host, "无法解析请求 URL")
    if host.lower() in ("localhost", "127.0.0.1", "::1"):
        raise ToolPermissionError("不允许访问本机地址")
    try:
        ip = ipaddress.ip_address(host)
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved:
            raise ToolPermissionError("不允许访问内网/保留地址")
    except ValueError:
        pass
    return url


def _tool_error_formatter(
    ctx: RunContextWrapper[TestHubContext], exc: Exception
) -> str:
    """工具异常统一转成结构化 JSON 返回给模型。"""
    try:
        return json.dumps(
            {"success": False, "error": f"{type(exc).__name__}: {exc}"},
            ensure_ascii=False,
        )
    except Exception:
        return f"{type(exc).__name__}: {exc}"


def assistant_tool(name: str, *, permission: str = "none") -> Any:
    """声明式工具注册：生成 SDK FunctionTool 并放入 TOOL_REGISTRY。

    permission 仅作为元数据标注（project/session/none），实际校验由
    工具体内部的 require_* 完成，避免因参数名差异导致误判。
    """

    def deco(fn):
        try:
            tool = function_tool(
                name_override=name,
                failure_error_function=_tool_error_formatter,
            )(fn)
        except UserError:
            # 含 Dict/Any 参数时 strict schema 不支持，自动降级非 strict
            tool = function_tool(
                name_override=name,
                failure_error_function=_tool_error_formatter,
                strict_mode=False,
            )(fn)
        TOOL_REGISTRY[name] = tool
        fn._tool_name = name  # type: ignore[attr-defined]
        fn._tool_permission = permission  # type: ignore[attr-defined]
        return fn

    return deco


def get_registered_tools() -> List[FunctionTool]:
    """返回按注册顺序排列的 SDK 工具列表。"""
    return list(TOOL_REGISTRY.values())


def registered_tool_names() -> List[str]:
    return list(TOOL_REGISTRY.keys())


# ---------------------------------------------------------------
# 查询类工具
# ---------------------------------------------------------------


@assistant_tool("get_project_overview", permission="project")
def get_project_overview(ctx: RunContextWrapper[TestHubContext], project_id: int) -> Dict:
    """获取项目概览：接口数、用例数、执行状态等。

    Args:
        project_id: 项目ID。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project
    from apps.testcases.models import TestCase
    from apps.api_testing.models import ApiProject, ApiRequest, TestExecution

    project = Project.objects.get(id=project_id)
    testcase_count = TestCase.objects.filter(project_id=project_id).count()

    api_projects = list(ApiProject.objects.filter(main_project=project)[:3])
    api_count = 0
    collection_count = 0
    recent_executions = []
    for ap in api_projects:
        api_count += ApiRequest.objects.filter(collection__project=ap).count()
        api_count += ApiRequest.objects.filter(collection__isnull=True).count()
        collection_count += ap.collections.count()
        execs = TestExecution.objects.filter(
            test_suite__project=ap
        ).order_by("-created_at")[:3].values(
            "id", "status", "total_requests", "passed_requests",
            "failed_requests", "created_at",
        )
        recent_executions.extend(list(execs))

    ui_count = 0
    ui_suite_count = 0
    try:
        from apps.ui_automation.models import UiProject, TestScript as UIScript, TestSuite as UITestSuite

        for up in UiProject.objects.filter(main_project=project)[:5]:
            ui_count += UIScript.objects.filter(project=up).count()
            ui_suite_count += UITestSuite.objects.filter(project=up).count()
    except Exception:
        pass

    midscene_count = 0
    midscene_device_count = 0
    try:
        from apps.ui_automation.models import MidsceneProject, MidsceneCase, MidsceneDevice

        for m in MidsceneProject.objects.filter(main_project=project)[:5]:
            midscene_count += MidsceneCase.objects.filter(project=m).count()
        midscene_device_count = MidsceneDevice.objects.filter(status="online").count()
    except Exception:
        pass

    return {
        "project_name": project.name,
        "project_status": project.status,
        "project_description": project.description or "",
        "has_knowledge_base": bool(project.knowledge_base),
        "modules": {
            "testcases": {"count": testcase_count},
            "api_testing": {
                "api_project_count": len(api_projects),
                "collection_count": collection_count,
                "api_count": api_count,
            },
            "ui_automation": {
                "script_count": ui_count,
                "suite_count": ui_suite_count,
            },
            "midscene": {
                "case_count": midscene_count,
                "online_devices": midscene_device_count,
            },
        },
        "recent_executions": [
            {
                "id": e["id"],
                "status": e["status"],
                "passed": e["passed_requests"],
                "total": e["total_requests"],
            }
            for e in recent_executions[:5]
        ],
    }


@assistant_tool("search_apis", permission="project")
def search_apis(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    keyword: str = "",
) -> Dict:
    """搜索API接口，按名称/URL/方法匹配。

    Args:
        project_id: 项目ID。
        keyword: 关键词，匹配名称/URL/方法。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project
    from apps.api_testing.models import ApiProject, ApiRequest
    from django.db.models import Q

    main_project = Project.objects.get(id=project_id)
    api_project_ids = list(
        ApiProject.objects.filter(main_project=main_project).values_list("id", flat=True)[:5]
    )
    queryset = ApiRequest.objects.filter(
        Q(collection__project_id__in=api_project_ids) | Q(collection__isnull=True)
    )
    if keyword:
        queryset = queryset.filter(
            Q(name__icontains=keyword)
            | Q(url__icontains=keyword)
            | Q(method__iexact=keyword)
        )
    total = queryset.count()
    results = [
        {"id": r.id, "name": r.name, "method": r.method, "url": r.url[:120]}
        for r in queryset[:10]
    ]
    resp = {"total": total, "count": len(results), "results": results}
    if total > 10:
        resp["hint"] = f"共 {total} 条结果，仅展示前 10 条。可缩小关键词精确查找。"
    return resp


@assistant_tool("get_api_detail", permission="project")
def get_api_detail(ctx: RunContextWrapper[TestHubContext], request_id: int) -> Dict:
    """获取API接口完整信息：URL、方法、请求头、参数、请求体、断言。

    Args:
        request_id: 接口ID。
    """
    req = require_api_access(ctx, request_id)
    headers, headers_truncated = _slim_json(req.headers or {})
    params, params_truncated = _slim_json(req.params or {})
    body, body_truncated = _slim_json(req.body or {})
    assertions, assertions_truncated = _slim_json(req.assertions or [])
    return {
        "id": req.id,
        "name": req.name,
        "method": req.method,
        "url": req.url,
        "headers": headers,
        "headers_truncated": headers_truncated,
        "params": params,
        "params_truncated": params_truncated,
        "body": body,
        "body_truncated": body_truncated,
        "assertions": assertions,
        "assertions_truncated": assertions_truncated,
        "collection": req.collection.name if req.collection else None,
        "collection_id": req.collection_id,
        "project": req.collection.project.name if req.collection and req.collection.project else None,
    }


@assistant_tool("search_testcases", permission="project")
def search_testcases(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    keyword: str = "",
    page: int = 1,
    page_size: int = 10,
) -> Dict:
    """搜索测试用例，按标题/描述关键词匹配，支持分页。

    Args:
        project_id: 项目ID。
        keyword: 搜索关键词。
        page: 页码，从1开始。
        page_size: 每页条数，最大200。
    """
    require_project(ctx, project_id)
    from apps.testcases.models import TestCase
    from django.db.models import Q

    queryset = TestCase.objects.filter(project_id=project_id)
    if keyword:
        queryset = queryset.filter(
            Q(title__icontains=keyword) | Q(description__icontains=keyword)
        )
    total = queryset.count()
    page = max(1, int(page))
    page_size = max(1, min(200, int(page_size)))
    start = (page - 1) * page_size
    results = [
        {"id": tc.id, "title": tc.title, "priority": tc.priority, "status": tc.status}
        for tc in queryset.order_by("-created_at")[start : start + page_size]
    ]
    total_pages = (total + page_size - 1) // page_size
    resp = {
        "total": total,
        "count": len(results),
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "results": results,
    }
    if page < total_pages:
        resp["hint"] = (
            f"共 {total} 条，当前第 {page}/{total_pages} 页。"
            "可翻页继续查看，或缩小关键词精确查找。"
        )
    return resp


@assistant_tool("analyze_testcases", permission="project")
def analyze_testcases(
    ctx: RunContextWrapper[TestHubContext], project_id: int
) -> Dict:
    """统计分析项目全部测试用例：状态/优先级/类型/模块分布、质量缺口、重复标题、执行状态。

    适合对大批量用例做整体分析；在工具内完成聚合，
    只返回统计摘要与少量样本，避免全量用例占用模型上下文。

    Args:
        project_id: 项目ID。
    """
    require_project(ctx, project_id)
    from django.db.models import Count
    from apps.testcases.models import TestCase, TestCaseStep

    qs = TestCase.objects.filter(project_id=project_id)
    total = qs.count()
    if total == 0:
        return {"total": 0, "analysis": "该项目暂无测试用例"}

    def _group_counts(field: str) -> Dict[str, int]:
        return {
            row[field]: row["c"]
            for row in qs.values(field).annotate(c=Count("id")).order_by("-c")
            if row[field]
        }

    module_rows = (
        qs.exclude(function_module__isnull=True)
        .values("function_module__name")
        .annotate(c=Count("id"))
        .order_by("-c")[:10]
    )
    by_module = {row["function_module__name"]: row["c"] for row in module_rows}

    missing_description = qs.filter(description="").count()
    missing_expected = qs.filter(expected_result="").count()
    with_steps = set(
        TestCaseStep.objects.filter(testcase__project_id=project_id)
        .values_list("testcase_id", flat=True)
    )
    no_steps = 0
    for tc in qs.only("id", "steps").iterator(chunk_size=500):
        if not (tc.steps or "").strip() and tc.id not in with_steps:
            no_steps += 1

    duplicates = [
        {"title": row["title"][:100], "count": row["c"]}
        for row in qs.values("title")
        .annotate(c=Count("id"))
        .filter(c__gt=1)
        .order_by("-c")[:10]
    ]

    exec_status = {
        row["execution_status"] or "not_run": row["c"]
        for row in qs.values("execution_status").annotate(c=Count("id"))
    }
    recent = [
        {
            "id": tc.id,
            "title": tc.title[:100],
            "status": tc.status,
            "priority": tc.priority,
        }
        for tc in qs.order_by("-created_at")[:5]
    ]

    return {
        "total": total,
        "by_status": _group_counts("status"),
        "by_priority": _group_counts("priority"),
        "by_type": _group_counts("test_type"),
        "by_module": by_module,
        "execution_status": exec_status,
        "quality_gaps": {
            "missing_description": missing_description,
            "missing_expected_result": missing_expected,
            "missing_steps": no_steps,
        },
        "duplicate_titles": duplicates,
        "recent": recent,
        "hint": (
            "已按数据库聚合统计，如需查看具体用例可配合 "
            "search_testcases（分页）或 get_testcase_detail（单条详情）。"
        ),
    }


# ---------------------------------------------------------------
# 创建/更新/删除类工具
# ---------------------------------------------------------------


@assistant_tool("create_api_test", permission="project")
def create_api_test(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    name: str,
    method: str,
    url: str,
    collection_id: Optional[int] = None,
    headers: Optional[Dict] = None,
    params: Optional[Dict] = None,
    body: Optional[Dict] = None,
    assertions: Optional[List] = None,
) -> Dict:
    """创建API接口测试，含URL、方法、请求头、参数、请求体、断言。

    Args:
        project_id: 项目ID。
        name: 接口名称。
        method: GET/POST/PUT/DELETE/PATCH。
        url: 请求URL。
        collection_id: 集合ID（可选）。
        headers: 请求头JSON。
        params: URL参数JSON。
        body: 请求体JSON。
        assertions: 断言规则列表。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project
    from apps.api_testing.models import ApiProject, ApiCollection, ApiRequest

    main_project = Project.objects.get(id=project_id)
    api_project = ApiProject.objects.filter(main_project=main_project).first()
    if not api_project:
        api_project = ApiProject.objects.create(
            name=main_project.name,
            project_type="HTTP",
            status="IN_PROGRESS",
            owner=ctx.context.user or main_project.owner,
            main_project=main_project,
        )

    collection = None
    if collection_id:
        collection = ApiCollection.objects.filter(
            id=collection_id, project=api_project
        ).first()

    request = ApiRequest.objects.create(
        collection=collection,
        name=name,
        method=method or "GET",
        url=url,
        headers=headers or {},
        params=params or {},
        body=body or {},
        assertions=assertions or [],
        created_by=ctx.context.user,
    )
    return {
        "success": True,
        "id": request.id,
        "name": request.name,
        "method": request.method,
        "url": request.url,
        "api_project": api_project.name,
    }


@assistant_tool("update_api_test", permission="project")
def update_api_test(
    ctx: RunContextWrapper[TestHubContext],
    request_id: int,
    name: Optional[str] = None,
    method: Optional[str] = None,
    url: Optional[str] = None,
    headers: Optional[Dict] = None,
    params: Optional[Dict] = None,
    body: Optional[Dict] = None,
    assertions: Optional[List] = None,
) -> Dict:
    """修改API接口测试的字段（仅传需修改的）。

    Args:
        request_id: 接口ID。
        name: 新名称。
        method: GET/POST/PUT/DELETE/PATCH。
        url: 新URL。
        headers: 请求头JSON。
        params: URL参数JSON。
        body: 请求体JSON。
        assertions: 断言规则（整体替换）。
    """
    req = require_api_access(ctx, request_id)
    updatable = {
        "name": name,
        "method": method,
        "url": url,
        "headers": headers,
        "params": params,
        "body": body,
        "assertions": assertions,
    }
    changed = {}
    for k, v in updatable.items():
        if v is not None:
            old = getattr(req, k)
            setattr(req, k, v)
            changed[k] = {"old": str(old)[:100], "new": str(v)[:100]}
    if changed:
        req.save()
    return {
        "success": True,
        "id": req.id,
        "name": req.name,
        "changed_fields": list(changed.keys()),
    }


@assistant_tool("create_collection", permission="project")
def create_collection(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    name: str,
    description: str = "",
    parent_id: Optional[int] = None,
) -> Dict:
    """创建接口集合（文件夹）。

    Args:
        project_id: 项目ID。
        name: 集合名称。
        description: 描述。
        parent_id: 父集合ID（嵌套用）。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project
    from apps.api_testing.models import ApiProject, ApiCollection

    main_project = Project.objects.get(id=project_id)
    api_project = ApiProject.objects.filter(main_project=main_project).first()
    if not api_project:
        api_project = ApiProject.objects.create(
            name=main_project.name,
            project_type="HTTP",
            status="IN_PROGRESS",
            owner=ctx.context.user or main_project.owner,
            main_project=main_project,
        )

    parent = None
    if parent_id:
        parent = ApiCollection.objects.filter(id=parent_id, project=api_project).first()
    collection = ApiCollection.objects.create(
        project=api_project,
        name=name,
        description=description,
        parent=parent,
    )
    return {
        "success": True,
        "id": collection.id,
        "name": collection.name,
        "api_project_id": api_project.id,
        "parent": parent.name if parent else None,
    }


@assistant_tool("create_testcase", permission="project")
def create_testcase(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    title: str,
    priority: str = "medium",
    status: str = "draft",
    description: str = "",
    steps: Optional[Any] = None,
) -> Dict:
    """创建测试用例，含标题、步骤、优先级。

    Args:
        project_id: 项目ID。
        title: 用例标题。
        priority: low/medium/high/critical。
        status: draft/active/deprecated。
        description: 描述/前置条件。
        steps: 测试步骤文本或步骤对象数组。
    """
    require_project(ctx, project_id)
    from apps.testcases.models import TestCase

    testcase = TestCase.objects.create(
        project_id=project_id,
        title=title,
        priority=priority or "medium",
        status=status or "draft",
        description=description or "",
        author=ctx.context.user,
    )
    step_count = 0
    if steps:
        if isinstance(steps, str):
            testcase.steps = steps
            testcase.save()
            step_count = 1
        elif isinstance(steps, list):
            from apps.testcases.models import TestCaseStep

            for i, step_data in enumerate(steps):
                TestCaseStep.objects.create(
                    testcase=testcase,
                    step_number=i + 1,
                    action=step_data.get("step", step_data.get("action", "")),
                    expected=step_data.get("expected", ""),
                )
            step_count = len(steps)
    return {
        "success": True,
        "id": testcase.id,
        "title": testcase.title,
        "priority": testcase.priority,
        "step_count": step_count,
    }


@assistant_tool("get_testcase_detail", permission="project")
def get_testcase_detail(
    ctx: RunContextWrapper[TestHubContext], testcase_id: int
) -> Dict:
    """查看测试用例详情：标题、描述、前置条件、步骤、预期结果、执行状态等。

    Args:
        testcase_id: 用例ID。
    """
    tc = require_testcase_access(ctx, testcase_id)
    steps = [
        {"step": s.step_number, "action": s.action, "expected": s.expected}
        for s in tc.step_details.all()
    ]
    steps_json, steps_truncated = _slim_json(steps)
    description, description_truncated = _slim_json(tc.description or "")
    preconditions, preconditions_truncated = _slim_json(tc.preconditions or "")
    expected_result, expected_result_truncated = _slim_json(tc.expected_result or "")
    return {
        "id": tc.id,
        "title": tc.title,
        "description": description,
        "description_truncated": description_truncated,
        "preconditions": preconditions,
        "preconditions_truncated": preconditions_truncated,
        "priority": tc.priority,
        "status": tc.status,
        "execution_status": tc.execution_status,
        "test_type": tc.test_type or "",
        "expected_result": expected_result,
        "expected_result_truncated": expected_result_truncated,
        "step_count": len(steps),
        "steps": steps_json,
        "steps_truncated": steps_truncated,
        "tags": tc.tags or [],
        "function_module": tc.function_module.name if tc.function_module else None,
        "versions": [v.name for v in tc.versions.all()],
        "author": getattr(tc.author, "username", ""),
        "assignee": getattr(tc.assignee, "username", "") if tc.assignee else None,
        "created_at": tc.created_at.isoformat() if tc.created_at else None,
        "updated_at": tc.updated_at.isoformat() if tc.updated_at else None,
    }


@assistant_tool("update_testcase", permission="project")
def update_testcase(
    ctx: RunContextWrapper[TestHubContext],
    testcase_id: int,
    title: Optional[str] = None,
    description: Optional[str] = None,
    priority: Optional[str] = None,
    status: Optional[str] = None,
    preconditions: Optional[str] = None,
    expected_result: Optional[str] = None,
) -> Dict:
    """修改测试用例字段（只传需修改的）。

    Args:
        testcase_id: 用例ID。
        title: 新标题。
        description: 新描述。
        priority: low/medium/high/critical。
        status: draft/active/deprecated。
        preconditions: 前置条件。
        expected_result: 预期结果。
    """
    tc = require_testcase_access(ctx, testcase_id)
    updatable = {
        "title": title,
        "description": description,
        "priority": priority,
        "status": status,
        "preconditions": preconditions,
        "expected_result": expected_result,
    }
    changed = {}
    for k, v in updatable.items():
        if v is not None:
            setattr(tc, k, v)
            changed[k] = v
    if changed:
        tc.save()
    return {
        "success": True,
        "id": tc.id,
        "title": tc.title,
        "changed": list(changed.keys()),
    }


@assistant_tool("delete_testcase", permission="project")
def delete_testcase(
    ctx: RunContextWrapper[TestHubContext], testcase_id: int
) -> Dict:
    """删除测试用例（不可恢复）。

    Args:
        testcase_id: 用例ID。
    """
    tc = require_testcase_access(ctx, testcase_id)
    title = tc.title
    tc.delete()
    return {"success": True, "deleted": title}


@assistant_tool("update_knowledge_base", permission="project")
def update_knowledge_base(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    content: str,
    mode: str = "append",
) -> Dict:
    """更新项目知识库内容。

    Args:
        project_id: 项目ID。
        content: 知识库内容。
        mode: append=追加, replace=替换。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project

    project = Project.objects.get(id=project_id)
    if mode == "replace":
        project.knowledge_base = content
    else:
        existing = project.knowledge_base or ""
        project.knowledge_base = existing + "\n\n" + content
    project.save()
    return {
        "success": True,
        "project": project.name,
        "knowledge_base_length": len(project.knowledge_base),
    }


def _require_suite_access(
    ctx: RunContextWrapper[TestHubContext], test_suite: Any
) -> Any:
    """校验测试套件所属项目访问权。"""
    api_project = test_suite.project
    if api_project.main_project_id:
        require_project(ctx, api_project.main_project_id)
    else:
        _require(
            _module_access(ctx.context.user, api_project),
            f"无权访问接口项目 {api_project.name}",
        )
    return test_suite


@assistant_tool("create_test_suite", permission="project")
def create_test_suite(
    ctx: RunContextWrapper[TestHubContext],
    project_id: int,
    name: str,
    description: str = "",
    environment_id: Optional[int] = None,
) -> Dict:
    """创建自动化测试套件，用于把接口组织成执行用例批次。

    Args:
        project_id: TestHub主项目ID。
        name: 套件名称。
        description: 套件描述。
        environment_id: 执行环境ID（可选，属于该接口项目）。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project
    from apps.api_testing.models import ApiProject, Environment, TestSuite

    main_project = Project.objects.get(id=project_id)
    api_project = ApiProject.objects.filter(main_project=main_project).first()
    if not api_project:
        return {
            "success": False,
            "error": "该项目下还没有接口测试项目，请先用 create_api_test 创建接口。",
        }

    environment = None
    if environment_id:
        environment = Environment.objects.filter(
            id=environment_id, project=api_project
        ).first()
        if not environment:
            return {"success": False, "error": f"环境 {environment_id} 不属于该接口项目"}

    suite = TestSuite.objects.create(
        project=api_project,
        name=name,
        description=description,
        environment=environment,
        created_by=ctx.context.user,
    )
    return {
        "success": True,
        "id": suite.id,
        "name": suite.name,
        "api_project": api_project.name,
    }


@assistant_tool("list_test_suites", permission="project")
def list_test_suites(
    ctx: RunContextWrapper[TestHubContext], project_id: int
) -> Dict:
    """列出项目下的自动化测试套件。

    Args:
        project_id: TestHub主项目ID。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project
    from apps.api_testing.models import ApiProject, TestSuite

    main_project = Project.objects.get(id=project_id)
    api_project = ApiProject.objects.filter(main_project=main_project).first()
    if not api_project:
        return {"total": 0, "count": 0, "results": []}
    suites = TestSuite.objects.filter(project=api_project)
    results = [
        {
            "id": s.id,
            "name": s.name,
            "description": (s.description or "")[:100],
            "environment": s.environment.name if s.environment else None,
            "case_count": s.testsuiterequest_set.count(),
            "created_at": s.created_at.isoformat() if s.created_at else None,
        }
        for s in suites.order_by("-created_at")[:20]
    ]
    resp = {"total": suites.count(), "count": len(results), "results": results}
    if suites.count() > 20:
        resp["hint"] = f"共 {suites.count()} 个套件，仅展示前 20 个。"
    return resp


@assistant_tool("add_suite_request", permission="project")
def add_suite_request(
    ctx: RunContextWrapper[TestHubContext],
    test_suite_id: int,
    request_id: int,
    name: str = "",
    description: str = "",
    order: Optional[int] = None,
    params: Optional[Dict] = None,
    headers: Optional[Dict] = None,
    body: Optional[Dict] = None,
    assertions: Optional[List] = None,
    extract_rules: Optional[List] = None,
    enabled: bool = True,
) -> Dict:
    """把接口添加为测试套件的执行用例，可覆盖参数/请求头/请求体/断言并配置变量提取。

    同一个接口可添加多条用例（参数或断言不同）。

    Args:
        test_suite_id: 测试套件ID。
        request_id: 接口ID。
        name: 用例名称，默认用接口名。
        description: 用例描述。
        order: 执行顺序，默认追加到末尾。
        params: 参数覆盖，如 {"userid": "123"}。
        headers: 请求头覆盖。
        body: 请求体覆盖。
        assertions: 断言规则列表（该用例专属，执行时与接口断言合并）。
        extract_rules: 响应变量提取规则列表。
        enabled: 是否启用该用例。
    """
    from apps.api_testing.models import TestSuite, TestSuiteRequest

    suite = _require_suite_access(ctx, TestSuite.objects.get(id=test_suite_id))
    api_request = require_api_access(ctx, request_id)
    if order is None:
        order = suite.testsuiterequest_set.count()
    suite_request = TestSuiteRequest.objects.create(
        test_suite=suite,
        request=api_request,
        name=name or "",
        description=description,
        order=order,
        assertions=assertions or [],
        params=params or {},
        headers=headers or {},
        body=body if body is not None else {},
        extract_rules=extract_rules or [],
        enabled=enabled,
    )
    return {
        "success": True,
        "id": suite_request.id,
        "case_name": suite_request.name or api_request.name,
        "request_id": api_request.id,
        "method": api_request.method,
        "url": api_request.url,
        "suite": suite.name,
        "order": suite_request.order,
        "existing_cases_in_suite": suite.testsuiterequest_set.count(),
    }


@assistant_tool("list_suite_requests", permission="project")
def list_suite_requests(
    ctx: RunContextWrapper[TestHubContext], test_suite_id: int
) -> Dict:
    """列出测试套件内的执行用例。

    Args:
        test_suite_id: 测试套件ID。
    """
    from apps.api_testing.models import TestSuite

    suite = _require_suite_access(ctx, TestSuite.objects.get(id=test_suite_id))
    cases = suite.testsuiterequest_set.select_related("request").order_by("order")
    results = [
        {
            "id": c.id,
            "name": c.name or c.request.name,
            "request_id": c.request_id,
            "method": c.request.method,
            "url": c.request.url[:120],
            "order": c.order,
            "enabled": c.enabled,
            "has_params_override": bool(c.params),
            "has_headers_override": bool(c.headers),
            "has_body_override": c.body is not None and bool(c.body),
            "assertion_count": len(c.assertions or []),
            "extract_rule_count": len(c.extract_rules or []),
        }
        for c in cases
    ]
    return {
        "suite_id": suite.id,
        "suite_name": suite.name,
        "total": len(results),
        "results": results,
    }


@assistant_tool("update_suite_request", permission="project")
def update_suite_request(
    ctx: RunContextWrapper[TestHubContext],
    suite_request_id: int,
    name: Optional[str] = None,
    description: Optional[str] = None,
    order: Optional[int] = None,
    params: Optional[Dict] = None,
    headers: Optional[Dict] = None,
    body: Optional[Dict] = None,
    assertions: Optional[List] = None,
    extract_rules: Optional[List] = None,
    enabled: Optional[bool] = None,
) -> Dict:
    """修改套件执行用例字段（只传需修改的；JSON字段传空对象/数组表示清空）。

    Args:
        suite_request_id: 套件用例ID。
        name: 用例名称。
        description: 用例描述。
        order: 执行顺序。
        params: 参数覆盖。
        headers: 请求头覆盖。
        body: 请求体覆盖。
        assertions: 断言规则。
        extract_rules: 响应变量提取规则。
        enabled: 是否启用。
    """
    from apps.api_testing.models import TestSuiteRequest

    suite_request = TestSuiteRequest.objects.select_related("test_suite", "request").get(
        id=suite_request_id
    )
    _require_suite_access(ctx, suite_request.test_suite)
    changed = []
    if name is not None:
        suite_request.name = name
        changed.append("name")
    if description is not None:
        suite_request.description = description
        changed.append("description")
    if order is not None:
        suite_request.order = order
        changed.append("order")
    if params is not None:
        suite_request.params = params
        changed.append("params")
    if headers is not None:
        suite_request.headers = headers
        changed.append("headers")
    if body is not None:
        suite_request.body = body
        changed.append("body")
    if assertions is not None:
        suite_request.assertions = assertions
        changed.append("assertions")
    if extract_rules is not None:
        suite_request.extract_rules = extract_rules
        changed.append("extract_rules")
    if enabled is not None:
        suite_request.enabled = enabled
        changed.append("enabled")
    if changed:
        suite_request.save()
    return {
        "success": True,
        "id": suite_request.id,
        "case_name": suite_request.name or suite_request.request.name,
        "changed_fields": changed,
    }


@assistant_tool("delete_suite_request", permission="project")
def delete_suite_request(
    ctx: RunContextWrapper[TestHubContext], suite_request_id: int
) -> Dict:
    """从测试套件中删除执行用例（不可恢复）。

    Args:
        suite_request_id: 套件用例ID。
    """
    from apps.api_testing.models import TestSuiteRequest

    suite_request = TestSuiteRequest.objects.select_related("test_suite").get(
        id=suite_request_id
    )
    _require_suite_access(ctx, suite_request.test_suite)
    case_name = suite_request.name or suite_request.request.name
    suite_request.delete()
    return {"success": True, "deleted": case_name}


@assistant_tool("execute_test_suite", permission="project")
def execute_test_suite(
    ctx: RunContextWrapper[TestHubContext], test_suite_id: int
) -> Dict:
    """执行测试套件，按顺序运行套件内启用的用例，返回执行统计与精简结果。

    Args:
        test_suite_id: 测试套件ID。
    """
    from apps.api_testing.models import TestSuite
    from apps.api_testing.services import execute_suite

    suite = _require_suite_access(ctx, TestSuite.objects.get(id=test_suite_id))
    result = execute_suite(suite, suite.environment, ctx.context.user)
    if not result.get("success"):
        return {"success": False, "error": result.get("error", "执行失败")}

    slim_results = [
        {
            "case_name": c.get("case_name") or c.get("name"),
            "method": c.get("method"),
            "url": (c.get("url") or "")[:120],
            "status_code": c.get("status_code"),
            "response_time_ms": c.get("response_time"),
            "passed": c.get("passed"),
            "error": (c.get("error") or "")[:200],
        }
        for c in result.get("results", [])
    ]
    return {
        "success": True,
        "suite": suite.name,
        "execution_id": result.get("execution_id"),
        "total_count": result.get("total_count"),
        "passed_count": result.get("passed_count"),
        "failed_count": result.get("failed_count"),
        "results": slim_results,
        "hint": "完整执行详情见执行记录/报告页；需要单用例细节可用 get_api_detail 或执行记录查询。",
    }


# ---------------------------------------------------------------
# 执行类工具
# ---------------------------------------------------------------


_AGENT_BROWSER_ACTIONS = {
    "open", "snapshot", "click", "dblclick", "focus", "fill", "type", "press",
    "keydown", "keyup", "hover", "check", "uncheck", "select", "scroll",
    "scrollintoview", "drag", "upload", "get", "is", "screenshot", "pdf",
    "record", "wait", "mouse", "find", "set", "cookies", "storage",
    "network", "back", "forward", "reload", "close",
}


def _find_agent_browser_binary() -> Optional[str]:
    """查找 agent-browser CLI（npm 全局安装，兼容 .cmd/.exe）。"""
    for name in ("agent-browser", "agent-browser.cmd", "agent-browser.exe"):
        p = shutil.which(name)
        if p:
            return p
    return None


@assistant_tool("agent_browser", permission="none")
def agent_browser(
    ctx: RunContextWrapper[TestHubContext],
    command: str,
) -> Dict:
    """执行 Agent Browser 浏览器自动化命令（headless Chrome），返回页面状态。

    配合 Agent Browser 技能使用，典型流程：
    1. open <url> 打开页面
    2. snapshot -i 获取可交互元素 refs（如 @e1）
    3. click <ref> / fill <ref> <text> / press <key> 操作元素
    4. get text <ref> / get url / get title 读取页面信息

    command 是 agent-browser 子命令（不含前缀），如：
    "open https://example.com"、"snapshot -i"、"click @e1"、
    "fill @e2 用户名"、"press Enter"、"get url"、"wait --text 加载完成"。

    Args:
        command: agent-browser 命令行。
    """
    cmd = (command or "").strip()
    if not cmd:
        return {"success": False, "error": "请提供 agent-browser 命令，如 open <url>"}
    action = cmd.split(maxsplit=1)[0].lower()
    if action not in _AGENT_BROWSER_ACTIONS:
        return {"success": False, "error": f"不支持的 agent-browser 动作: {action}"}
    try:
        args = shlex.split(cmd)
    except ValueError as e:
        return {"success": False, "error": f"命令参数解析失败: {e}"}

    binary = _find_agent_browser_binary()
    if not binary:
        return {
            "success": False,
            "error": "未找到 agent-browser CLI，请先执行 npm install -g agent-browser",
        }
    if os.name == "nt":
        full = ["cmd.exe", "/d", "/c", binary, *args]
    else:
        full = [binary, *args]
    try:
        proc = subprocess.run(
            full,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=60,
        )
    except subprocess.TimeoutExpired:
        return {"success": False, "error": "agent-browser 执行超时（60s）"}
    except Exception as e:
        return {"success": False, "error": f"agent-browser 执行失败: {e}"}

    output = proc.stdout or ""
    stderr = proc.stderr or ""
    truncated = len(output) > 4000
    return {
        "success": proc.returncode == 0,
        "action": action,
        "returncode": proc.returncode,
        "output": output[:4000],
        "truncated": truncated,
        "stderr": (stderr or "")[:500],
        "hint": (
            "输出超过 4000 字符时已截断；继续操作请基于返回的 refs 调用 "
            "snapshot/click/fill 等，不要重复打开页面。"
        ),
    }


@assistant_tool("execute_api", permission="project")
def execute_api(
    ctx: RunContextWrapper[TestHubContext],
    request_id: int,
    environment_id: Optional[int] = None,
) -> Dict:
    """执行API请求并返回状态码、响应体、响应时间。

    Args:
        request_id: 接口ID。
        environment_id: 环境ID（可选）。
    """
    req = require_api_access(ctx, request_id)
    from apps.api_testing.models import Environment
    from apps.api_testing.utils import execute_api_request

    environment = None
    if environment_id:
        environment = Environment.objects.filter(id=environment_id).first()
    _validate_target_url(req.url, environment)

    result = execute_api_request(req, environment, ctx.context.user)
    raw_body = result.get("response_data", {}).get("body", "")
    try:
        body_str = json.dumps(
            json.loads(raw_body) if isinstance(raw_body, str) else raw_body,
            ensure_ascii=False,
            indent=2,
        )
    except Exception:
        body_str = str(raw_body) if raw_body else ""
    return {
        "status_code": result.get("status_code"),
        "response_time_ms": round(result.get("response_time", 0), 1),
        "response_body": (body_str or "")[:400],
        "body_truncated": len(body_str) > 400,
        "assertions": result.get("assertions_results"),
        "error": result.get("error_message", ""),
    }


# ---------------------------------------------------------------
# 项目发现工具
# ---------------------------------------------------------------


@assistant_tool("list_api_projects", permission="none")
def list_api_projects(
    ctx: RunContextWrapper[TestHubContext], keyword: str = ""
) -> Dict:
    """列出当前用户可访问的API测试项目。

    Args:
        keyword: 按名称筛选。
    """
    from apps.api_testing.models import ApiProject
    from django.db.models import Q

    user = ctx.context.user
    queryset = ApiProject.objects.all() if getattr(user, "is_superuser", False) else ApiProject.objects.filter(
        Q(owner=user) | Q(members=user)
    )
    if keyword:
        queryset = queryset.filter(name__icontains=keyword)
    total = queryset.count()
    results = [
        {"id": p.id, "name": p.name, "status": p.status}
        for p in queryset[:10]
    ]
    resp = {"total": total, "count": len(results), "results": results}
    if total > 10:
        resp["hint"] = f"共 {total} 个 API 项目，仅展示前 10 个。可用 keyword 筛选。"
    return resp


@assistant_tool("list_midscene_projects", permission="none")
def list_midscene_projects(
    ctx: RunContextWrapper[TestHubContext], keyword: str = ""
) -> Dict:
    """列出当前用户可访问的Midscene（AI智能模式）项目。

    Args:
        keyword: 按名称筛选。
    """
    from apps.ui_automation.models import MidsceneProject
    from django.db.models import Q

    user = ctx.context.user
    queryset = (
        MidsceneProject.objects.all()
        if getattr(user, "is_superuser", False)
        else MidsceneProject.objects.filter(Q(owner=user) | Q(members=user))
    )
    if keyword:
        queryset = queryset.filter(name__icontains=keyword)
    total = queryset.count()
    results = [
        {"id": p.id, "name": p.name, "description": (p.description or "")[:80]}
        for p in queryset[:10]
    ]
    resp = {"total": total, "count": len(results), "results": results}
    if total > 10:
        resp["hint"] = f"共 {total} 个 Midscene 项目，仅展示前 10 个。"
    return resp


@assistant_tool("list_midscene_cases", permission="project")
def list_midscene_cases(
    ctx: RunContextWrapper[TestHubContext],
    midscene_project_id: int,
    keyword: str = "",
) -> Dict:
    """列出Midscene项目下的AI智能模式用例。

    Args:
        midscene_project_id: Midscene项目ID。
        keyword: 按名称筛选。
    """
    require_midscene_access(ctx, midscene_project_id)
    from apps.ui_automation.models import MidsceneCase

    queryset = MidsceneCase.objects.filter(project_id=midscene_project_id)
    if keyword:
        queryset = queryset.filter(name__icontains=keyword)
    total = queryset.count()
    results = [
        {
            "id": c.id,
            "name": c.name,
            "description": (c.description or "")[:100],
            "ai_prompt": (c.ai_prompt or "")[:100],
            "max_steps": c.max_steps,
        }
        for c in queryset[:10]
    ]
    resp = {"total": total, "count": len(results), "results": results}
    if total > 10:
        resp["hint"] = f"共 {total} 个 Midscene 用例，仅展示前 10 个。"
    return resp


@assistant_tool("update_midscene_case", permission="project")
def update_midscene_case(
    ctx: RunContextWrapper[TestHubContext],
    case_id: int,
    name: Optional[str] = None,
    ai_prompt: Optional[str] = None,
    description: Optional[str] = None,
    max_steps: Optional[int] = None,
    append_step: Optional[str] = None,
) -> Dict:
    """修改Midscene用例（ai_prompt 每行一个步骤，换行分隔，禁止用→连接）。

    Args:
        case_id: 用例ID。
        name: 新名称。
        ai_prompt: AI Prompt（换行分隔步骤）。
        description: 新描述。
        max_steps: 最大执行步数。
        append_step: 追加步骤（自动换行）。
    """
    from apps.ui_automation.models import MidsceneCase

    case = MidsceneCase.objects.get(id=case_id)
    if case.project_id:
        require_midscene_access(ctx, case.project_id)
    changed = {}

    if append_step:
        existing = (case.ai_prompt or "").strip()
        case.ai_prompt = existing + "\n" + append_step.strip() if existing else append_step.strip()
        changed["ai_prompt"] = "appended"

    updatable = {"name": name, "ai_prompt": ai_prompt, "description": description, "max_steps": max_steps}
    for k, v in updatable.items():
        if v is not None:
            setattr(case, k, v)
            changed[k] = {"old": None, "new": v}
    if changed:
        case.save()
    return {
        "success": True,
        "id": case.id,
        "name": case.name,
        "ai_prompt": case.ai_prompt,
        "changed_fields": list(changed.keys()),
    }


# ---------------------------------------------------------------
# 知识库工具
# ---------------------------------------------------------------


@assistant_tool("read_knowledge_base", permission="project")
def read_knowledge_base(
    ctx: RunContextWrapper[TestHubContext], project_id: int
) -> Dict:
    """读取项目知识库内容。

    Args:
        project_id: 项目ID。
    """
    require_project(ctx, project_id)
    from apps.projects.models import Project

    project = Project.objects.get(id=project_id)
    if not project.knowledge_base:
        return {"content": "", "message": "该项目未配置知识库"}
    content = project.knowledge_base[:3000]
    return {
        "project_name": project.name,
        "content": content,
        "truncated": len(project.knowledge_base) > 3000,
        "length": len(project.knowledge_base),
    }


# ---------------------------------------------------------------
# 文档解析工具
# ---------------------------------------------------------------


@assistant_tool("parse_swagger", permission="none")
def parse_swagger(
    ctx: RunContextWrapper[TestHubContext],
    url: Optional[str] = None,
    content: Optional[str] = None,
) -> Dict:
    """解析OpenAPI/Swagger文档，提取接口列表。

    Args:
        url: 文档URL。
        content: 文档JSON内容。
    """
    import requests as sync_requests

    try:
        spec = None
        if url:
            _validate_target_url(url)
            resp = sync_requests.get(url, timeout=30)
            spec = resp.json()
        elif content:
            spec = json.loads(content)
        else:
            return {"error": "请提供url或content参数"}

        endpoints = []
        tags_map = {}
        for tag in spec.get("tags", []):
            tags_map[tag["name"]] = tag.get("description", "")

        for path, methods in spec.get("paths", {}).items():
            for method, detail in methods.items():
                if method in ("get", "post", "put", "delete", "patch", "options", "head"):
                    tag = detail.get("tags", ["默认"])[0] if detail.get("tags") else "默认"
                    endpoints.append(
                        {
                            "path": path,
                            "method": method.upper(),
                            "summary": detail.get("summary", ""),
                            "tag": tag,
                            "parameters": detail.get("parameters", []),
                            "request_body": bool(detail.get("requestBody")),
                            "responses": list(detail.get("responses", {}).keys()),
                        }
                    )

        groups = {}
        for ep in endpoints:
            groups.setdefault(ep["tag"], []).append(ep)
        slim_endpoints = [
            {
                "path": ep["path"],
                "method": ep["method"],
                "summary": ep["summary"][:80],
                "tag": ep["tag"],
            }
            for ep in endpoints[:10]
        ]
        result = {
            "total_endpoints": len(endpoints),
            "groups": {tag: len(eps) for tag, eps in groups.items()},
            "sample_count": len(slim_endpoints),
            "endpoints": slim_endpoints,
        }
        if len(endpoints) > 10:
            result["hint"] = f"共 {len(endpoints)} 个接口，仅展示前 10 条概要。可按 tag 分组创建或指定具体接口。"
        return result
    except Exception as e:
        return {"error": str(e)}


@assistant_tool("parse_yapi", permission="none")
def parse_yapi(
    ctx: RunContextWrapper[TestHubContext],
    file_path: Optional[str] = None,
    content: Optional[str] = None,
    limit: Optional[int] = None,
) -> Dict:
    """解析YApi导出的JSON，提取接口列表。

    Args:
        file_path: 会话内JSON文件路径。
        content: JSON内容。
        limit: 最多解析的接口数，默认解析全部。
    """
    try:
        data = None
        if file_path:
            fp = _safe_session_path(ctx, file_path)
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
        elif content:
            data = json.loads(content)
        else:
            return {"error": "请提供file_path或content"}

        items = []
        if isinstance(data, list):
            for group in data:
                if isinstance(group, dict) and "list" in group:
                    items.extend(group["list"])
                else:
                    items.append(group)
        elif isinstance(data, dict):
            if "data" in data:
                inner = data["data"]
                if isinstance(inner, dict) and "list" in inner:
                    items = inner["list"]
                elif isinstance(inner, list):
                    items = inner
            elif "list" in data:
                items = data["list"]

        if not items:
            return {"error": "未识别的YApi格式，请确认是YApi导出的JSON"}

        # 接口过多时默认只返回概要，避免大文档全量字段撑爆上下文；
        # 显式传 limit 时按用户意图返回详情。
        summary_mode = limit is None and len(items) > 50
        endpoints = []
        items_to_parse = items if limit is None else items[:limit]
        for item in items_to_parse:
            path = item.get("path", "") or (item.get("query_path") or {}).get("path", "")
            ep = {
                "title": item.get("title", ""),
                "path": path,
                "method": (item.get("method") or "GET").upper(),
                "desc": (item.get("desc") or "")[:100],
            }
            if summary_mode:
                endpoints.append(ep)
                continue
            req_query = item.get("req_query", [])
            if req_query:
                ep["params"] = [
                    {
                        "name": q.get("name", ""),
                        "required": q.get("required") == "1",
                        "desc": q.get("desc", ""),
                    }
                    for q in req_query
                ]
            body = item.get("req_body_other", "") or item.get("req_body_form", [])
            if isinstance(body, str) and body:
                try:
                    body = json.loads(body)
                except Exception:
                    pass
            ep["req_body"] = str(body)[:200] if body else ""
            res = item.get("res_body", "")
            if isinstance(res, str) and res:
                try:
                    res = json.loads(res)
                except Exception:
                    pass
            ep["res_body"] = str(res)[:200] if res else ""
            endpoints.append(ep)

        result = {"total": len(items), "count": len(endpoints), "endpoints": endpoints}
        if summary_mode:
            result["summary_mode"] = True
            result["hint"] = (
                f"共 {len(items)} 个接口，为避免占用上下文仅返回概要"
                "（方法/路径/标题）。如需详情可按版本或分组指定接口，"
                "或传 limit 获取前 N 个的完整字段。"
            )
        if limit is not None and len(items) > limit:
            result["hint"] = (
                f"共 {len(items)} 个接口，limit={limit}，仅解析前 {limit} 条。"
                "如需全部接口请不传limit。"
            )
        return result
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------------
# 会话文件工具
# ---------------------------------------------------------------


@assistant_tool("list_session_files", permission="session")
def list_session_files(ctx: RunContextWrapper[TestHubContext]) -> Dict:
    """列出当前用户所有会话中已上传的文件。"""
    from apps.assistant.models import AgentFile

    user = ctx.context.user
    files = AgentFile.objects.filter(source="upload")
    if not getattr(user, "is_superuser", False):
        files = files.filter(session__user=user)
    files = files.order_by("-created_at")[:20]
    results = [
        {
            "id": f.id,
            "file_name": f.file_name,
            "file_path": f.file_path,
            "file_size": f.file_size,
        }
        for f in files
    ]
    return {
        "count": len(results),
        "files": results,
        "hint": "用 read_session_file(file_path=...) 读内容",
    }


@assistant_tool("read_session_file", permission="session")
def read_session_file(
    ctx: RunContextWrapper[TestHubContext], file_path: str
) -> Dict:
    """读取会话上传文件内容（JSON/YAML 用本工具，PDF/Word/Excel 用 simple_doc_parser）。

    Args:
        file_path: 文件路径（必须位于当前用户会话上传目录内）。
    """
    fp = _safe_session_path(ctx, file_path)
    try:
        with open(fp, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        return {
            "file": fp,
            "size": len(content),
            "content": content[:3000],
            "truncated": len(content) > 3000,
        }
    except FileNotFoundError:
        return {"error": f"文件不存在: {fp}"}
    except Exception as e:
        return {"error": str(e)}


@assistant_tool("simple_doc_parser", permission="session")
def simple_doc_parser(
    ctx: RunContextWrapper[TestHubContext], file_path: str
) -> Dict:
    """读取会话内文档文件内容（PDF/Word/Excel/TXT/HTML/CSV/JSON/YAML）。

    Args:
        file_path: 文件路径（必须位于当前用户会话上传目录内）。
    """
    fp = _safe_session_path(ctx, file_path)
    ext = os.path.splitext(fp)[1].lower()
    try:
        if ext == ".pdf":
            from pypdf import PdfReader

            text = "\n".join((page.extract_text() or "") for page in PdfReader(fp).pages)
        elif ext == ".docx":
            from docx import Document

            text = "\n".join(p.text for p in Document(fp).paragraphs)
        elif ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            wb = load_workbook(fp, read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join("" if c is None else str(c) for c in row))
            text = "\n".join(rows)
        elif ext in (".html", ".htm"):
            from bs4 import BeautifulSoup

            text = BeautifulSoup(
                Path(fp).read_text(encoding="utf-8", errors="replace"),
                "html.parser",
            ).get_text("\n")
        else:
            text = Path(fp).read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return {"success": False, "error": f"解析失败（{ext or '未知格式'}）: {e}"}
    return {
        "file": fp,
        "type": ext or "text",
        "size": os.path.getsize(fp),
        "content": text[:3000],
        "truncated": len(text) > 3000,
    }
